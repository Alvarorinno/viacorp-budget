"""
Script para reimportar datos desde Excel hacia eventos_2026.json
Uso: python3 scripts/import_excel.py <ruta_excel> <contraseña>
"""
import sys
import json
import msoffcrypto
import io
from openpyxl import load_workbook

def import_excel(path, password):
    with open(path, 'rb') as f:
        office_file = msoffcrypto.OfficeFile(f)
        office_file.load_key(password=password)
        decrypted = io.BytesIO()
        office_file.decrypt(decrypted)

    decrypted.seek(0)
    wb = load_workbook(decrypted, data_only=True)
    ws = wb['2026']

    rows = []
    for row in ws.iter_rows(min_row=3, max_row=200, values_only=True):
        b = row[1]
        if b is None or (isinstance(b, str) and b in ('MES', 'SUMA', 'ESTIMACION')):
            continue
        if row[2] is None:
            continue
        rows.append({
            'estimacion': row[1],
            'cliente': row[2],
            'descripcion': row[3],
            'presupuesto': row[4],
            'costo': row[5],
            'mb': row[6],
            'factura': str(row[7]) if row[7] is not None else None,
            'fecha_facturacion': str(row[8]) if row[8] is not None else None,
            'mes_evento': row[9],
            'por_cobrar': str(row[11]) if row[11] is not None else None,
            'pagado': str(row[12]) if row[12] is not None else None,
        })

    output = 'scripts/eventos_2026.json'
    with open(output, 'w', encoding='utf-8') as f:
        json.dump(rows, f, ensure_ascii=False, default=str)
    print(f'Importados {len(rows)} registros → {output}')

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print('Uso: python3 scripts/import_excel.py <archivo.xlsx> <contraseña>')
        sys.exit(1)
    import_excel(sys.argv[1], sys.argv[2])
